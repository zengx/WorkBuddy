#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
create_blank_xlsx.py

Create a blank xlsx workbook on disk for the sheet-generation pipeline.

Usage:
    python3 create_blank_xlsx.py <output_path> <title> [sheet_names_csv]

Arguments:
    output_path       Absolute or relative path to the .xlsx file to create.
                      Must end with ".xlsx" (case-insensitive); ".xls" and
                      ".xlsm" are intentionally rejected here even though the
                      downstream edit pipeline accepts them — we always emit
                      modern xlsx for new workbooks.
                      Parent directories are created if missing.
                      Existing file at the path is overwritten silently
                      (no --force flag, no prompt).
    title             Workbook title written into OOXML core metadata
                      (docProps/core.xml dc:title). Free-form string.
    sheet_names_csv   (optional) Comma-separated worksheet names.
                      Default: "Sheet1".
                      Each name must:
                        - have length 1..31
                        - not contain any of: [ ] : * ? / \
                        - not start or end with a single quote
                      Duplicate names (case-insensitive) are rejected.

Behavior:
    - On success, prints the absolute path of the created file to stdout.
    - Always overwrites an existing file at output_path; the caller is
      expected to choose a stable target path per generation task.
    - Output is a standard OOXML xlsx, directly consumable by the existing
      edit pipeline (resolve_local_excel → open_file → set_cell_range ...).

Exit codes:
    0  success
    2  argument / validation error
    3  filesystem / write error
    4  dependency installation failed

Notes:
    - openpyxl is auto-installed via `pip install --quiet openpyxl` if missing,
      following the same pattern as the PPT skill's python-pptx usage.
    - Use python3 explicitly; do not rely on shell `python` aliases.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

# -------- openpyxl auto-install (strategy A) --------
try:
    from openpyxl import Workbook  # type: ignore
except ImportError:
    print("[create_blank_xlsx] openpyxl not found, installing ...", file=sys.stderr)
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl>=3.1.0"]
        )
    except subprocess.CalledProcessError as e:
        print(
            f"[create_blank_xlsx] failed to install openpyxl: {e}",
            file=sys.stderr,
        )
        sys.exit(4)
    try:
        from openpyxl import Workbook  # type: ignore  # noqa: E402
    except ImportError as e:
        print(
            f"[create_blank_xlsx] openpyxl still unavailable after install: {e}",
            file=sys.stderr,
        )
        sys.exit(4)

# -------- Excel sheet name rules --------
# Forbidden chars per Excel spec: [ ] : * ? / \
_FORBIDDEN_CHARS = set("[]:*?/\\")
_MAX_NAME_LEN = 31


def _validate_sheet_name(name: str) -> str | None:
    """Return None if OK, else an error message."""
    if not name:
        return "sheet name is empty"
    if len(name) > _MAX_NAME_LEN:
        return f"sheet name '{name}' exceeds {_MAX_NAME_LEN} chars"
    bad = [c for c in name if c in _FORBIDDEN_CHARS]
    if bad:
        return f"sheet name '{name}' contains forbidden chars: {''.join(sorted(set(bad)))}"
    if name.startswith("'") or name.endswith("'"):
        return f"sheet name '{name}' must not start or end with a single quote"
    return None


def _parse_sheet_names(csv: str) -> list[str]:
    raw = [s.strip() for s in csv.split(",")]
    names = [s for s in raw if s]
    if not names:
        raise ValueError("sheet_names_csv parsed to empty list")
    seen: dict[str, str] = {}
    for n in names:
        err = _validate_sheet_name(n)
        if err:
            raise ValueError(err)
        key = n.lower()
        if key in seen:
            raise ValueError(
                f"duplicate sheet name (case-insensitive): '{n}' vs '{seen[key]}'"
            )
        seen[key] = n
    return names


def _print_usage_and_exit(code: int = 2) -> None:
    print(__doc__, file=sys.stderr)
    sys.exit(code)


def main(argv: list[str]) -> int:
    if len(argv) < 3 or len(argv) > 4:
        _print_usage_and_exit(2)

    output_path_arg = argv[1]
    title = argv[2]
    sheet_names_csv = argv[3] if len(argv) == 4 else "Sheet1"

    # validate title (lenient: any non-empty string)
    if not isinstance(title, str) or not title.strip():
        print("[create_blank_xlsx] title must be a non-empty string", file=sys.stderr)
        return 2

    # validate output path
    output_path = Path(output_path_arg).expanduser()
    if output_path.suffix.lower() != ".xlsx":
        print(
            f"[create_blank_xlsx] output_path must end with .xlsx, got: {output_path}",
            file=sys.stderr,
        )
        return 2

    # validate sheet names
    try:
        sheet_names = _parse_sheet_names(sheet_names_csv)
    except ValueError as e:
        print(f"[create_blank_xlsx] invalid sheet_names_csv: {e}", file=sys.stderr)
        return 2

    # ensure parent dir
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        print(
            f"[create_blank_xlsx] failed to create parent dir {output_path.parent}: {e}",
            file=sys.stderr,
        )
        return 3

    # build workbook
    wb = Workbook()
    # Workbook() ships with one default sheet named 'Sheet'; remove it then add ours
    default = wb.active
    wb.remove(default)
    for name in sheet_names:
        wb.create_sheet(title=name)
    # write OOXML core metadata title
    try:
        wb.properties.title = title
    except Exception as e:  # pragma: no cover - openpyxl quirks
        print(
            f"[create_blank_xlsx] warning: failed to set workbook title metadata: {e}",
            file=sys.stderr,
        )

    # save (overwrite if exists)
    try:
        wb.save(str(output_path))
    except OSError as e:
        print(
            f"[create_blank_xlsx] failed to write {output_path}: {e}",
            file=sys.stderr,
        )
        return 3

    abs_path = str(output_path.resolve())
    print(abs_path)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
